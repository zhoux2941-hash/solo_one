from datetime import datetime

from planner import PACE_TYPES
from heart_rate import calculate_heart_rate_zones, calculate_hr_for_pace, get_hr_zone_color

WEEKDAYS = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']

def export_to_excel(plan, training_days, file_path):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("请先安装openpyxl: pip install openpyxl")
    
    wb = openpyxl.Workbook()
    
    info_ws = wb.active
    info_ws.title = "计划信息"
    
    title_font = Font(size=16, bold=True)
    header_font = Font(size=12, bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    info_ws['A1'] = "跑步训练计划"
    info_ws['A1'].font = title_font
    info_ws.merge_cells('A1:F1')
    info_ws['A1'].alignment = center_alignment
    
    max_hr = plan.get('max_heart_rate', 0)
    
    info_data = [
        ["计划名称", plan['name']],
        ["训练目标", plan['goal']],
        ["初始周跑量", f"{plan['current_weekly_distance']} 公里"],
        ["每周训练天数", f"{plan['training_days']} 天"],
        ["开始日期", plan['start_date']],
        ["生成时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    if max_hr > 0:
        info_data.append(["最大心率", f"{max_hr} bpm"])
    
    for row, (key, value) in enumerate(info_data, start=3):
        info_ws[f'A{row}'] = key
        info_ws[f'A{row}'].font = header_font
        info_ws[f'B{row}'] = value
        info_ws[f'A{row}'].border = thin_border
        info_ws[f'B{row}'].border = thin_border
    
    info_ws.column_dimensions['A'].width = 15
    info_ws.column_dimensions['B'].width = 40
    
    pace_colors = {
        'easy': '90EE90',
        'tempo': 'FFD700',
        'interval': 'FF6347',
        'long': '87CEFA',
        'rest': 'D3D3D3'
    }
    
    overview_ws = wb.create_sheet("训练总览")
    
    overview_ws['A1'] = "8周训练计划总览"
    overview_ws['A1'].font = title_font
    overview_ws.merge_cells('A1:I1')
    overview_ws['A1'].alignment = center_alignment
    
    headers = ['周次'] + WEEKDAYS
    for col, header in enumerate(headers, start=1):
        cell = overview_ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(size=11, bold=True, color='FFFFFF')
    
    week_data = {}
    for day in training_days:
        week = day['week_number']
        day_idx = day['day_of_week'] - 1
        if week not in week_data:
            week_data[week] = {}
        week_data[week][day_idx] = day
    
    for week in range(1, 9):
        row = week + 3
        week_cell = overview_ws.cell(row=row, column=1, value=f"第 {week} 周")
        week_cell.font = Font(bold=True)
        week_cell.alignment = center_alignment
        week_cell.border = thin_border
        
        week_days = week_data.get(week, {})
        for day_idx in range(7):
            col = day_idx + 2
            day = week_days.get(day_idx)
            
            if day:
                if day['pace_type'] == 'rest':
                    display_text = "休息"
                else:
                    display_text = f"{day['distance']}km\n{PACE_TYPES.get(day['pace_type'], day['pace_type'])}"
            else:
                display_text = "-"
            
            cell = overview_ws.cell(row=row, column=col, value=display_text)
            cell.alignment = center_alignment
            cell.border = thin_border
            
            if day and day['pace_type'] in pace_colors:
                cell.fill = PatternFill(
                    start_color=pace_colors[day['pace_type']],
                    end_color=pace_colors[day['pace_type']],
                    fill_type='solid'
                )
    
    for col in range(1, 10):
        overview_ws.column_dimensions[get_column_letter(col)].width = 14
    
    for week in range(1, 9):
        week_ws = wb.create_sheet(f"第{week}周")
        
        week_ws['A1'] = f"第 {week} 周训练计划"
        week_ws['A1'].font = title_font
        week_ws.merge_cells('A1:G1')
        week_ws['A1'].alignment = center_alignment
        
        for col, day_name in enumerate(WEEKDAYS, start=1):
            cell = week_ws.cell(row=3, column=col, value=day_name)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(size=11, bold=True, color='FFFFFF')
        
        week_days = week_data.get(week, {})
        
        row_titles = ['距离(km)', '配速类型']
        if max_hr > 0:
            row_titles.append('建议心率')
        row_titles.append('备注')
        
        for row_idx, row_title in enumerate(row_titles, start=4):
            row_title_cell = week_ws.cell(row=row_idx, column=1, value=row_title)
            row_title_cell.font = Font(bold=True)
            row_title_cell.alignment = center_alignment
            row_title_cell.border = thin_border
        
        hr_row_idx = 6 if max_hr > 0 else 5
        notes_row_idx = hr_row_idx + 1
        
        for day_idx in range(7):
            col = day_idx + 1
            day = week_days.get(day_idx)
            
            if day:
                distance = day['distance'] if day['pace_type'] != 'rest' else '-'
                pace = PACE_TYPES.get(day['pace_type'], day['pace_type'])
                notes = day.get('notes', '')
                
                distance_cell = week_ws.cell(row=4, column=col, value=distance)
                pace_cell = week_ws.cell(row=5, column=col, value=pace)
                
                cells_to_format = [distance_cell, pace_cell]
                
                if max_hr > 0:
                    if day['pace_type'] == 'rest':
                        hr_text = '休息'
                        hr_color = pace_colors['rest']
                    else:
                        hr_info = calculate_hr_for_pace(day['pace_type'], max_hr)
                        if hr_info:
                            hr_text = f"{hr_info['range']} bpm\n({hr_info['zone_name']})"
                            hr_color = get_hr_zone_color(hr_info['zone'])
                        else:
                            hr_text = '-'
                            hr_color = 'FFFFFF'
                    
                    hr_cell = week_ws.cell(row=hr_row_idx, column=col, value=hr_text)
                    hr_cell.alignment = center_alignment
                    hr_cell.border = thin_border
                    hr_cell.fill = PatternFill(
                        start_color=hr_color,
                        end_color=hr_color,
                        fill_type='solid'
                    )
                    cells_to_format.append(hr_cell)
                
                notes_cell = week_ws.cell(row=notes_row_idx, column=col, value=notes)
                cells_to_format.append(notes_cell)
                
                for cell in cells_to_format:
                    cell.alignment = center_alignment
                    cell.border = thin_border
                    
                    if day['pace_type'] in pace_colors and cell is not (hr_cell if max_hr > 0 else None):
                        if cell is notes_cell or cell is distance_cell or cell is pace_cell:
                            cell.fill = PatternFill(
                                start_color=pace_colors[day['pace_type']],
                                end_color=pace_colors[day['pace_type']],
                                fill_type='solid'
                            )
            else:
                end_row = notes_row_idx + 1
                for row in range(4, end_row):
                    cell = week_ws.cell(row=row, column=col, value='-')
                    cell.alignment = center_alignment
                    cell.border = thin_border
        
        for col in range(1, 8):
            week_ws.column_dimensions[get_column_letter(col)].width = 16
    
    stats_ws = wb.create_sheet("统计信息")
    
    stats_ws['A1'] = "训练计划统计"
    stats_ws['A1'].font = title_font
    stats_ws.merge_cells('A1:D1')
    stats_ws['A1'].alignment = center_alignment
    
    total_distance = sum(d['distance'] for d in training_days)
    pace_counts = {}
    for day in training_days:
        pace = day['pace_type']
        if pace not in pace_counts:
            pace_counts[pace] = 0
        pace_counts[pace] += 1
    
    stats_data = [
        ["统计项", "数值"],
        ["总训练距离", f"{total_distance:.1f} 公里"],
        ["平均每周距离", f"{total_distance / 8:.1f} 公里"],
        ["训练天数", f"{sum(1 for d in training_days if d['pace_type'] != 'rest')} 天"],
        ["休息日数", f"{pace_counts.get('rest', 0)} 天"],
        ["轻松跑天数", f"{pace_counts.get('easy', 0)} 天"],
        ["节奏跑天数", f"{pace_counts.get('tempo', 0)} 天"],
        ["间歇跑天数", f"{pace_counts.get('interval', 0)} 天"],
        ["长距离跑天数", f"{pace_counts.get('long', 0)} 天"]
    ]
    
    for row, (key, value) in enumerate(stats_data, start=3):
        key_cell = stats_ws.cell(row=row, column=1, value=key)
        value_cell = stats_ws.cell(row=row, column=2, value=value)
        
        if row == 3:
            key_cell.font = header_font
            value_cell.font = header_font
            key_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            value_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            key_cell.font = Font(size=11, bold=True, color='FFFFFF')
            value_cell.font = Font(size=11, bold=True, color='FFFFFF')
        elif key in pace_counts:
            if key in pace_colors:
                key_cell.fill = PatternFill(
                    start_color=pace_colors.get(key, 'FFFFFF'),
                    end_color=pace_colors.get(key, 'FFFFFF'),
                    fill_type='solid'
                )
        
        key_cell.alignment = center_alignment
        value_cell.alignment = center_alignment
        key_cell.border = thin_border
        value_cell.border = thin_border
    
    stats_ws.column_dimensions['A'].width = 20
    stats_ws.column_dimensions['B'].width = 25
    
    if max_hr > 0:
        hr_zones = calculate_heart_rate_zones(max_hr)
        if hr_zones:
            hr_ws = wb.create_sheet("心率区间")
            
            hr_ws['A1'] = f"心率区间 (最大心率: {max_hr} bpm)"
            hr_ws['A1'].font = title_font
            hr_ws.merge_cells('A1:E1')
            hr_ws['A1'].alignment = center_alignment
            
            hr_headers = ['区间', '名称', '心率范围', 'RPE', '说明']
            for col, header in enumerate(hr_headers, start=1):
                cell = hr_ws.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.font = Font(size=11, bold=True, color='FFFFFF')
            
            zone_order = ['zone1', 'zone2', 'zone3', 'zone4', 'zone5']
            for row, zone_key in enumerate(zone_order, start=4):
                zone = hr_zones[zone_key]
                
                hr_ws.cell(row=row, column=1, value=zone_key).alignment = center_alignment
                hr_ws.cell(row=row, column=2, value=zone['name']).alignment = center_alignment
                hr_ws.cell(row=row, column=3, value=f"{zone['min_hr']}-{zone['max_hr']}").alignment = center_alignment
                hr_ws.cell(row=row, column=4, value=zone['rpe']).alignment = center_alignment
                hr_ws.cell(row=row, column=5, value=zone['description']).alignment = center_alignment
                
                hr_color = get_hr_zone_color(zone_key)
                for col in range(1, 6):
                    cell = hr_ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.fill = PatternFill(
                        start_color=hr_color,
                        end_color=hr_color,
                        fill_type='solid'
                    )
            
            hr_ws.column_dimensions['A'].width = 8
            hr_ws.column_dimensions['B'].width = 12
            hr_ws.column_dimensions['C'].width = 15
            hr_ws.column_dimensions['D'].width = 8
            hr_ws.column_dimensions['E'].width = 40
    
    wb.save(file_path)


def export_to_pdf(plan, training_days, file_path):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    except ImportError:
        raise ImportError("请先安装reportlab: pip install reportlab")
    
    doc = SimpleDocTemplate(
        file_path,
        pagesize=landscape(A4),
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=24,
        spaceAfter=20,
        textColor=colors.HexColor('#2E5090')
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=12,
        textColor=colors.HexColor('#2E5090')
    )
    
    pace_colors = {
        'easy': colors.HexColor('#90EE90'),
        'tempo': colors.HexColor('#FFD700'),
        'interval': colors.HexColor('#FF6347'),
        'long': colors.HexColor('#87CEFA'),
        'rest': colors.HexColor('#D3D3D3')
    }
    
    max_hr = plan.get('max_heart_rate', 0)
    
    story = []
    
    story.append(Paragraph("🏃 跑步训练计划", title_style))
    story.append(Spacer(1, 20))
    
    info_data = [
        ['计划名称', plan['name']],
        ['训练目标', plan['goal']],
        ['初始周跑量', f"{plan['current_weekly_distance']} 公里"],
        ['每周训练天数', f"{plan['training_days']} 天"],
        ['开始日期', plan['start_date']],
        ['生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    if max_hr > 0:
        info_data.append(['最大心率', f"{max_hr} bpm"])
    
    info_table = Table(info_data, colWidths=[5*cm, 15*cm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('#E8F0FE')),
    ]))
    story.append(info_table)
    
    story.append(Spacer(1, 20))
    
    story.append(Paragraph("配速类型说明", heading_style))
    legend_data = [[]]
    for pace_type, display_name in PACE_TYPES.items():
        legend_data[0].append(display_name)
    
    legend_table = Table(legend_data, colWidths=[3*cm] * len(PACE_TYPES))
    legend_style = [
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (0, 0), pace_colors['easy']),
        ('BACKGROUND', (1, 0), (1, 0), pace_colors['tempo']),
        ('BACKGROUND', (2, 0), (2, 0), pace_colors['interval']),
        ('BACKGROUND', (3, 0), (3, 0), pace_colors['rest']),
        ('BACKGROUND', (4, 0), (4, 0), pace_colors['long']),
    ]
    legend_table.setStyle(legend_style)
    story.append(legend_table)
    
    story.append(Spacer(1, 20))
    
    week_data = {}
    for day in training_days:
        week = day['week_number']
        day_idx = day['day_of_week'] - 1
        if week not in week_data:
            week_data[week] = {}
        week_data[week][day_idx] = day
    
    for week in range(1, 9):
        if week > 1:
            story.append(PageBreak())
        
        story.append(Paragraph(f"第 {week} 周训练计划", heading_style))
        
        table_data = [['周次'] + WEEKDAYS]
        
        week_days = week_data.get(week, {})
        row_data = [f"第{week}周"]
        
        for day_idx in range(7):
            day = week_days.get(day_idx)
            if day:
                if day['pace_type'] == 'rest':
                    row_data.append("休息")
                else:
                    row_data.append(f"{day['distance']}km\n{PACE_TYPES.get(day['pace_type'], day['pace_type'])}")
            else:
                row_data.append("-")
        
        table_data.append(row_data)
        
        week_table = Table(table_data, colWidths=[2*cm] + [3.5*cm]*7, rowHeights=[1*cm, 1.5*cm])
        
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (0, 1), 'Helvetica-Bold'),
        ]
        
        for day_idx in range(7):
            day = week_days.get(day_idx)
            if day and day['pace_type'] in pace_colors:
                table_style.append(('BACKGROUND', (day_idx + 1, 1), (day_idx + 1, 1), pace_colors[day['pace_type']]))
        
        week_table.setStyle(table_style)
        story.append(week_table)
        
        story.append(Spacer(1, 15))
        
        if max_hr > 0:
            detail_headers = ['日期', '距离(km)', '配速类型', '建议心率', '备注']
            detail_col_widths = [3.5*cm, 2.5*cm, 3.5*cm, 5*cm, 7*cm]
        else:
            detail_headers = ['日期', '距离(km)', '配速类型', '备注']
            detail_col_widths = [4*cm, 3*cm, 4*cm, 10*cm]
        
        detail_data = [detail_headers]
        for day_idx in range(7):
            day = week_days.get(day_idx)
            if day:
                row = [
                    day['date'],
                    str(day['distance']) if day['pace_type'] != 'rest' else '-',
                    PACE_TYPES.get(day['pace_type'], day['pace_type']),
                ]
                
                if max_hr > 0:
                    if day['pace_type'] == 'rest':
                        row.append('休息')
                    else:
                        hr_info = calculate_hr_for_pace(day['pace_type'], max_hr)
                        if hr_info:
                            row.append(f"{hr_info['range']}\n({hr_info['zone_name']})")
                        else:
                            row.append('-')
                
                row.append(day.get('notes', ''))
                detail_data.append(row)
        
        if len(detail_data) > 1:
            detail_table = Table(detail_data, colWidths=detail_col_widths)
            detail_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]
            
            detail_table.setStyle(detail_style)
            story.append(detail_table)
    
    story.append(PageBreak())
    
    story.append(Paragraph("训练计划统计", heading_style))
    
    total_distance = sum(d['distance'] for d in training_days)
    pace_counts = {}
    for day in training_days:
        pace = day['pace_type']
        if pace not in pace_counts:
            pace_counts[pace] = 0
        pace_counts[pace] += 1
    
    stats_data = [
        ['统计项', '数值'],
        ['总训练距离', f"{total_distance:.1f} 公里"],
        ['平均每周距离', f"{total_distance / 8:.1f} 公里"],
        ['训练天数', f"{sum(1 for d in training_days if d['pace_type'] != 'rest')} 天"],
        ['休息日数', f"{pace_counts.get('rest', 0)} 天"],
        ['轻松跑天数', f"{pace_counts.get('easy', 0)} 天"],
        ['节奏跑天数', f"{pace_counts.get('tempo', 0)} 天"],
        ['间歇跑天数', f"{pace_counts.get('interval', 0)} 天"],
        ['长距离跑天数', f"{pace_counts.get('long', 0)} 天"]
    ]
    
    stats_table = Table(stats_data, colWidths=[6*cm, 8*cm])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#E8F0FE')),
    ]))
    story.append(stats_table)
    
    if max_hr > 0:
        hr_zones = calculate_heart_rate_zones(max_hr)
        if hr_zones:
            story.append(Spacer(1, 30))
            story.append(Paragraph(f"心率区间 (最大心率: {max_hr} bpm)", heading_style))
            
            hr_data = [['区间', '名称', '心率范围', 'RPE', '说明']]
            zone_order = ['zone1', 'zone2', 'zone3', 'zone4', 'zone5']
            
            for zone_key in zone_order:
                zone = hr_zones[zone_key]
                hr_data.append([
                    zone_key,
                    zone['name'],
                    f"{zone['min_hr']}-{zone['max_hr']}",
                    zone['rpe'],
                    zone['description']
                ])
            
            hr_table = Table(hr_data, colWidths=[2*cm, 3*cm, 3*cm, 2*cm, 9*cm])
            hr_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]
            
            for row_idx in range(1, 6):
                zone_key = f'zone{row_idx}'
                color_hex = get_hr_zone_color(zone_key)
                hr_style.append(
                    ('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor(f'#{color_hex}'))
                )
            
            hr_table.setStyle(hr_style)
            story.append(hr_table)
    
    doc.build(story)
